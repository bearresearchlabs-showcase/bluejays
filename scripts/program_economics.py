from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Color constants ──
blue_font = Font(name='Arial', color='0000FF', size=11)
blue_font_bold = Font(name='Arial', color='0000FF', size=11, bold=True)
black_font = Font(name='Arial', color='000000', size=11)
black_bold = Font(name='Arial', color='000000', size=11, bold=True)
header_font = Font(name='Arial', color='FFFFFF', size=11, bold=True)
title_font = Font(name='Arial', color='000000', size=14, bold=True)
section_font = Font(name='Arial', color='000000', size=12, bold=True)
red_font = Font(name='Arial', color='FF0000', size=11, bold=True)
green_font = Font(name='Arial', color='008000', size=11, bold=True)

dark_fill = PatternFill('solid', fgColor='2F5496')
light_fill = PatternFill('solid', fgColor='D6E4F0')
yellow_fill = PatternFill('solid', fgColor='FFFF00')
light_gray = PatternFill('solid', fgColor='F2F2F2')
white_fill = PatternFill('solid', fgColor='FFFFFF')
red_fill = PatternFill('solid', fgColor='FFC7CE')
green_fill = PatternFill('solid', fgColor='C6EFCE')

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
bottom_border = Border(bottom=Side(style='medium'))

center = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')

dollar_fmt = '$#,##0'
dollar_fmt_neg = '$#,##0;($#,##0);"-"'
pct_fmt = '0.0%'
num_fmt = '#,##0'
decimal_fmt = '0.00'

def style_range(ws, row, cols, font=black_font, fill=white_fill, alignment=right_align, border=thin_border, number_format=None):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border
        if number_format:
            cell.number_format = number_format

def header_row(ws, row, data, fill=dark_fill, font=header_font):
    for i, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.font = font
        cell.fill = fill
        cell.alignment = center
        cell.border = thin_border

def data_row(ws, row, data, fonts=None, fills=None, formats=None):
    for i, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.font = (fonts[i-1] if fonts else black_font)
        cell.fill = (fills[i-1] if fills else white_fill)
        cell.alignment = left_align if i == 1 else right_align
        cell.border = thin_border
        if formats and formats[i-1]:
            cell.number_format = formats[i-1]

# ═══════════════════════════════════════════════════════
# SHEET 1: Program Scope & Unit Economics
# ═══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Program Economics'
ws1.sheet_properties.tabColor = '2F5496'

# Column widths
ws1.column_dimensions['A'].width = 42
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 22
ws1.column_dimensions['D'].width = 22
ws1.column_dimensions['E'].width = 22

# Title
ws1.merge_cells('A1:E1')
ws1.cell(row=1, column=1, value='Text-to-SQL Training Data Program Economics').font = title_font

# ── Section 1: Program Scope ──
ws1.merge_cells('A3:E3')
ws1.cell(row=3, column=1, value='PROGRAM SCOPE').font = section_font
ws1['A3'].fill = light_fill

header_row(ws1, 4, ['Parameter', 'Value', 'Unit', 'Notes', ''])

assumptions = [
    ['Total Databases', 100, 'databases', 'Client target scope', blue_font],
    ['Queries per Database', 30, 'queries/db', 'Current standard', blue_font],
    ['Total Query Pairs', None, 'NL-SQL pairs', '=B5*B6', None],
    ['Current Databases Delivered', 16, 'databases', 'db-1 through db-16', blue_font],
    ['Current Queries Delivered', 452, 'queries', 'Actual count', blue_font],
    ['Databases Requiring Rework', 6, 'databases', 'db-6,7,9,10,11,16 missing fields', blue_font],
    ['Remaining Databases', None, 'databases', '=B5-B8', None],
]

for i, row_data in enumerate(assumptions):
    r = 5 + i
    ws1.cell(row=r, column=1, value=row_data[0]).font = black_font
    ws1.cell(row=r, column=1).alignment = left_align
    ws1.cell(row=r, column=1).border = thin_border
    
    cell_b = ws1.cell(row=r, column=2)
    if row_data[1] is not None:
        cell_b.value = row_data[1]
        cell_b.font = row_data[4] if row_data[4] else black_font
    else:
        cell_b.value = row_data[5] if len(row_data) > 5 else None
        cell_b.font = black_font
    cell_b.alignment = right_align
    cell_b.border = thin_border
    cell_b.number_format = num_fmt
    
    ws1.cell(row=r, column=3, value=row_data[2]).font = black_font
    ws1.cell(row=r, column=3).alignment = center
    ws1.cell(row=r, column=3).border = thin_border
    
    ws1.cell(row=r, column=4, value=row_data[3]).font = Font(name='Arial', color='808080', size=10, italic=True)
    ws1.cell(row=r, column=4).alignment = left_align
    ws1.cell(row=r, column=4).border = thin_border

# Formulas for calculated rows
ws1['B7'] = '=B5*B6'
ws1['B7'].font = black_font
ws1['B11'] = '=B5-B8'
ws1['B11'].font = black_font

# ── Section 2: Per-Query Time Estimates ──
r = 13
ws1.merge_cells(f'A{r}:E{r}')
ws1.cell(row=r, column=1, value='PER-QUERY EFFORT BREAKDOWN (Hours)').font = section_font
ws1[f'A{r}'].fill = light_fill

header_row(ws1, r+1, ['Task Component', 'Min Hours', 'Likely Hours', 'Max Hours', 'Notes'])

tasks = [
    ['Schema comprehension & domain research', 0.25, 0.5, 1.0, 'Understand tables, relationships, business context'],
    ['SQL query design & validation', 0.5, 0.75, 1.5, 'Write, test, debug complex queries'],
    ['Business intent narrative (purpose)', 0.25, 0.5, 0.75, 'Why a stakeholder needs this analysis'],
    ['Use case documentation', 0.15, 0.25, 0.5, 'Who uses it, what decision it drives'],
    ['Business value articulation', 0.15, 0.25, 0.5, 'Revenue/cost/risk impact quantified'],
    ['Expected output specification', 0.1, 0.15, 0.25, 'Column descriptions, sample results'],
    ['Cross-query consistency review', 0.1, 0.15, 0.25, 'Ensure no overlap, progressive complexity'],
    ['Quality assurance & formatting', 0.15, 0.25, 0.5, 'JSON validation, style guide compliance'],
]

for i, t in enumerate(tasks):
    row = r + 2 + i
    ws1.cell(row=row, column=1, value=t[0]).font = black_font
    ws1.cell(row=row, column=1).alignment = left_align
    ws1.cell(row=row, column=1).border = thin_border
    for j in range(1, 4):
        cell = ws1.cell(row=row, column=j+1, value=t[j])
        cell.font = blue_font
        cell.alignment = right_align
        cell.border = thin_border
        cell.number_format = decimal_fmt
    ws1.cell(row=row, column=5, value=t[4]).font = Font(name='Arial', color='808080', size=10, italic=True)
    ws1.cell(row=row, column=5).alignment = left_align
    ws1.cell(row=row, column=5).border = thin_border

# Totals row
tot_row = r + 2 + len(tasks)
ws1.cell(row=tot_row, column=1, value='TOTAL PER QUERY').font = black_bold
ws1.cell(row=tot_row, column=1).fill = light_fill
ws1.cell(row=tot_row, column=1).border = thin_border
ws1.cell(row=tot_row, column=1).alignment = left_align
for col in [2, 3, 4]:
    cell = ws1.cell(row=tot_row, column=col)
    start = r + 2
    end = tot_row - 1
    col_letter = get_column_letter(col)
    cell.value = f'=SUM({col_letter}{start}:{col_letter}{end})'
    cell.font = black_bold
    cell.fill = light_fill
    cell.border = thin_border
    cell.alignment = right_align
    cell.number_format = decimal_fmt

# Per-database row
db_row = tot_row + 1
ws1.cell(row=db_row, column=1, value='TOTAL PER DATABASE (30 queries)').font = black_bold
ws1.cell(row=db_row, column=1).border = thin_border
ws1.cell(row=db_row, column=1).alignment = left_align
for col in [2, 3, 4]:
    col_letter = get_column_letter(col)
    cell = ws1.cell(row=db_row, column=col)
    cell.value = f'={col_letter}{tot_row}*B6'
    cell.font = black_bold
    cell.border = thin_border
    cell.alignment = right_align
    cell.number_format = decimal_fmt

# ── Section 3: Rate Comparison ──
rate_start = db_row + 2
ws1.merge_cells(f'A{rate_start}:E{rate_start}')
ws1.cell(row=rate_start, column=1, value='RATE ANALYSIS: STAFFING AGENCY vs CONSULTING').font = section_font
ws1[f'A{rate_start}'].fill = light_fill

header_row(ws1, rate_start+1, ['Metric', 'Staffing Agency', 'Consulting', 'Delta', 'Notes'])

# Rate inputs
ri = rate_start + 2
rows_rate = [
    ['Hourly Rate', 100, 475, None, 'Market rates for 10yr tech professional'],
    ['Effective Daily Rate (8 hrs)', None, None, None, '=hourly * 8'],
    ['Per-Query Cost (likely hrs)', None, None, None, '=rate * likely hrs/query'],
    ['Per-Database Cost (30 queries)', None, None, None, '=per-query * 30'],
    ['Full Program Cost (100 databases)', None, None, None, '=per-db * 100'],
    ['Current Contract Value per DB', 2000, 2000, None, 'Client offer: $2K/database'],
    ['Margin per Database (Staffing)', None, None, None, '=contract - cost'],
    ['Margin per Database (Consulting)', None, None, None, ''],
    ['Implied Hourly Rate at $2K/db', None, None, None, '=$2K / hours per db'],
    ['', None, None, None, ''],
    ['FULL PROGRAM COMPARISON', None, None, None, ''],
    ['Total Program Revenue (100 × $2K)', None, None, None, ''],
    ['Total Program Cost', None, None, None, ''],
    ['Program Profit / (Loss)', None, None, None, ''],
    ['Program Margin %', None, None, None, ''],
]

for i, rd in enumerate(rows_rate):
    row = ri + i
    ws1.cell(row=row, column=1, value=rd[0]).font = black_bold if rd[0].startswith('FULL') or rd[0].startswith('Program') or rd[0].startswith('Total Program') else black_font
    ws1.cell(row=row, column=1).alignment = left_align
    ws1.cell(row=row, column=1).border = thin_border

# Hardcoded inputs (blue)
ws1.cell(row=ri, column=2, value=100).font = blue_font
ws1.cell(row=ri, column=2).number_format = dollar_fmt
ws1.cell(row=ri, column=2).border = thin_border
ws1.cell(row=ri, column=2).alignment = right_align
ws1.cell(row=ri, column=3, value=475).font = blue_font
ws1.cell(row=ri, column=3).number_format = dollar_fmt
ws1.cell(row=ri, column=3).border = thin_border
ws1.cell(row=ri, column=3).alignment = right_align
ws1.cell(row=ri, column=4).value = f'=C{ri}-B{ri}'
ws1.cell(row=ri, column=4).font = black_font
ws1.cell(row=ri, column=4).number_format = dollar_fmt
ws1.cell(row=ri, column=4).border = thin_border
ws1.cell(row=ri, column=4).alignment = right_align

# Daily rate
dr = ri + 1
ws1.cell(row=dr, column=2).value = f'=B{ri}*8'
ws1.cell(row=dr, column=3).value = f'=C{ri}*8'
ws1.cell(row=dr, column=4).value = f'=C{dr}-B{dr}'
for c in [2,3,4]:
    ws1.cell(row=dr, column=c).font = black_font
    ws1.cell(row=dr, column=c).number_format = dollar_fmt
    ws1.cell(row=dr, column=c).border = thin_border
    ws1.cell(row=dr, column=c).alignment = right_align

# Per-query cost (likely hours)
pq = ri + 2
ws1.cell(row=pq, column=2).value = f'=B{ri}*C{tot_row}'
ws1.cell(row=pq, column=3).value = f'=C{ri}*C{tot_row}'
ws1.cell(row=pq, column=4).value = f'=C{pq}-B{pq}'
for c in [2,3,4]:
    ws1.cell(row=pq, column=c).font = black_font
    ws1.cell(row=pq, column=c).number_format = dollar_fmt
    ws1.cell(row=pq, column=c).border = thin_border
    ws1.cell(row=pq, column=c).alignment = right_align

# Per-database cost
pd_row = ri + 3
ws1.cell(row=pd_row, column=2).value = f'=B{pq}*B6'
ws1.cell(row=pd_row, column=3).value = f'=C{pq}*B6'
ws1.cell(row=pd_row, column=4).value = f'=C{pd_row}-B{pd_row}'
for c in [2,3,4]:
    ws1.cell(row=pd_row, column=c).font = black_font
    ws1.cell(row=pd_row, column=c).number_format = dollar_fmt
    ws1.cell(row=pd_row, column=c).border = thin_border
    ws1.cell(row=pd_row, column=c).alignment = right_align

# Full program cost
fp = ri + 4
ws1.cell(row=fp, column=2).value = f'=B{pd_row}*B5'
ws1.cell(row=fp, column=3).value = f'=C{pd_row}*B5'
ws1.cell(row=fp, column=4).value = f'=C{fp}-B{fp}'
for c in [2,3,4]:
    ws1.cell(row=fp, column=c).font = black_bold
    ws1.cell(row=fp, column=c).number_format = dollar_fmt
    ws1.cell(row=fp, column=c).border = thin_border
    ws1.cell(row=fp, column=c).alignment = right_align
    ws1.cell(row=fp, column=c).fill = light_fill

# Contract value per DB
cv = ri + 5
ws1.cell(row=cv, column=2, value=2000).font = blue_font
ws1.cell(row=cv, column=2).number_format = dollar_fmt
ws1.cell(row=cv, column=2).border = thin_border
ws1.cell(row=cv, column=2).alignment = right_align
ws1.cell(row=cv, column=2).fill = yellow_fill
ws1.cell(row=cv, column=3, value=2000).font = blue_font
ws1.cell(row=cv, column=3).number_format = dollar_fmt
ws1.cell(row=cv, column=3).border = thin_border
ws1.cell(row=cv, column=3).alignment = right_align
ws1.cell(row=cv, column=3).fill = yellow_fill

# Margin per DB staffing
ms = ri + 6
ws1.cell(row=ms, column=2).value = f'=B{cv}-B{pd_row}'
ws1.cell(row=ms, column=2).font = black_font
ws1.cell(row=ms, column=2).number_format = dollar_fmt_neg
ws1.cell(row=ms, column=2).border = thin_border
ws1.cell(row=ms, column=2).alignment = right_align

# Margin per DB consulting
mc = ri + 7
ws1.cell(row=mc, column=3).value = f'=C{cv}-C{pd_row}'
ws1.cell(row=mc, column=3).font = black_font
ws1.cell(row=mc, column=3).number_format = dollar_fmt_neg
ws1.cell(row=mc, column=3).border = thin_border
ws1.cell(row=mc, column=3).alignment = right_align

# Implied hourly rate
ihr = ri + 8
ws1.cell(row=ihr, column=2).value = f'=B{cv}/C{db_row}'
ws1.cell(row=ihr, column=2).font = red_font
ws1.cell(row=ihr, column=2).number_format = dollar_fmt
ws1.cell(row=ihr, column=2).border = thin_border
ws1.cell(row=ihr, column=2).alignment = right_align
ws1.cell(row=ihr, column=2).fill = red_fill

# Full program comparison section
fpc_start = ri + 10
ws1.cell(row=fpc_start, column=1).font = black_bold
ws1.cell(row=fpc_start, column=1).fill = light_fill
for c in [2,3,4]:
    ws1.cell(row=fpc_start, column=c).fill = light_fill
    ws1.cell(row=fpc_start, column=c).border = thin_border

# Total program revenue
tpr = ri + 11
ws1.cell(row=tpr, column=2).value = f'=B{cv}*B5'
ws1.cell(row=tpr, column=3).value = f'=C{cv}*B5'
for c in [2,3]:
    ws1.cell(row=tpr, column=c).font = black_font
    ws1.cell(row=tpr, column=c).number_format = dollar_fmt
    ws1.cell(row=tpr, column=c).border = thin_border
    ws1.cell(row=tpr, column=c).alignment = right_align

# Total program cost
tpc = ri + 12
ws1.cell(row=tpc, column=2).value = f'=B{fp}'
ws1.cell(row=tpc, column=3).value = f'=C{fp}'
for c in [2,3]:
    ws1.cell(row=tpc, column=c).font = black_font
    ws1.cell(row=tpc, column=c).number_format = dollar_fmt
    ws1.cell(row=tpc, column=c).border = thin_border
    ws1.cell(row=tpc, column=c).alignment = right_align

# Program profit/loss
ppl = ri + 13
ws1.cell(row=ppl, column=2).value = f'=B{tpr}-B{tpc}'
ws1.cell(row=ppl, column=3).value = f'=C{tpr}-C{tpc}'
for c in [2,3]:
    ws1.cell(row=ppl, column=c).font = black_bold
    ws1.cell(row=ppl, column=c).number_format = dollar_fmt_neg
    ws1.cell(row=ppl, column=c).border = thin_border
    ws1.cell(row=ppl, column=c).alignment = right_align

# Program margin %
pm = ri + 14
ws1.cell(row=pm, column=2).value = f'=IF(B{tpr}=0,0,B{ppl}/B{tpr})'
ws1.cell(row=pm, column=3).value = f'=IF(C{tpr}=0,0,C{ppl}/C{tpr})'
for c in [2,3]:
    ws1.cell(row=pm, column=c).font = black_bold
    ws1.cell(row=pm, column=c).number_format = pct_fmt
    ws1.cell(row=pm, column=c).border = thin_border
    ws1.cell(row=pm, column=c).alignment = right_align

# Notes column
for row in range(ri, pm+1):
    ws1.cell(row=row, column=5).border = thin_border
    ws1.cell(row=row, column=5).alignment = left_align

notes_map = {
    ri: 'Market rates for 10yr tech professional',
    ri+2: 'Using "likely" hours estimate',
    ri+5: 'Client offer: $2K per database',
    ri+8: 'What client is actually paying per hour',
}
for row, note in notes_map.items():
    ws1.cell(row=row, column=5, value=note).font = Font(name='Arial', color='808080', size=10, italic=True)

# ── Section 4: Industry Benchmarks ──
bench_start = pm + 2
ws1.merge_cells(f'A{bench_start}:E{bench_start}')
ws1.cell(row=bench_start, column=1, value='INDUSTRY BENCHMARKS: Text-to-SQL Annotation Pricing').font = section_font
ws1[f'A{bench_start}'].fill = light_fill

header_row(ws1, bench_start+1, ['Benchmark / Source', 'Price per NL-SQL Pair', 'Quality Tier', 'Complexity', 'Notes'])

benchmarks = [
    ['Crowdsource (MTurk basic)', 2, 'Low', 'Simple SELECT', 'No domain expertise'],
    ['Crowdsource (MTurk expert)', 8, 'Medium', 'Joins, subqueries', 'SQL-trained annotators'],
    ['Domain Expert Annotation', 15, 'High', 'Multi-table, CTEs', 'Industry + SQL knowledge'],
    ['Enterprise SME + Schema', 25, 'Premium', 'Full enterprise', 'Business context + intent'],
    ['AQ Current Contract', None, 'Premium', 'Full enterprise', '=$2K / 30 queries'],
    ['Break-Even at Staffing Rate', None, 'Premium', 'Full enterprise', '=staffing cost / 30'],
    ['Break-Even at Consulting Rate', None, 'Premium', 'Full enterprise', '=consulting cost / 30'],
]

for i, b in enumerate(benchmarks):
    row = bench_start + 2 + i
    ws1.cell(row=row, column=1, value=b[0]).font = black_font
    ws1.cell(row=row, column=1).alignment = left_align
    ws1.cell(row=row, column=1).border = thin_border
    
    cell_price = ws1.cell(row=row, column=2)
    if b[1] is not None:
        cell_price.value = b[1]
        cell_price.font = blue_font
    cell_price.number_format = dollar_fmt
    cell_price.border = thin_border
    cell_price.alignment = right_align
    
    ws1.cell(row=row, column=3, value=b[2]).font = black_font
    ws1.cell(row=row, column=3).alignment = center
    ws1.cell(row=row, column=3).border = thin_border
    
    ws1.cell(row=row, column=4, value=b[3]).font = black_font
    ws1.cell(row=row, column=4).alignment = center
    ws1.cell(row=row, column=4).border = thin_border
    
    ws1.cell(row=row, column=5, value=b[4]).font = Font(name='Arial', color='808080', size=10, italic=True)
    ws1.cell(row=row, column=5).alignment = left_align
    ws1.cell(row=row, column=5).border = thin_border

# AQ contract per pair
aq_row = bench_start + 6
ws1.cell(row=aq_row, column=2).value = f'=B{cv}/B6'
ws1.cell(row=aq_row, column=2).font = red_font
ws1.cell(row=aq_row, column=2).fill = red_fill

# Break-even staffing per pair
be_staff = bench_start + 7
ws1.cell(row=be_staff, column=2).value = f'=B{pd_row}/B6'
ws1.cell(row=be_staff, column=2).font = black_font

# Break-even consulting per pair
be_consult = bench_start + 8
ws1.cell(row=be_consult, column=2).value = f'=C{pd_row}/B6'
ws1.cell(row=be_consult, column=2).font = black_font

# ═══════════════════════════════════════════════════════
# SHEET 2: Deliverable Format Specification
# ═══════════════════════════════════════════════════════
ws2 = wb.create_sheet('Format Specification')
ws2.sheet_properties.tabColor = '008000'

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 55
ws2.column_dimensions['E'].width = 30

ws2.merge_cells('A1:E1')
ws2.cell(row=1, column=1, value='Recommended Deliverable JSON Schema — Text-to-SQL Training Format').font = title_font

ws2.merge_cells('A3:E3')
ws2.cell(row=3, column=1, value='QUERY-LEVEL FIELDS (per entry in "queries" array)').font = section_font
ws2['A3'].fill = light_fill

header_row(ws2, 4, ['Field Name', 'Type', 'Required', 'Description', 'Maps to Benchmark'])

fields = [
    ['number', 'integer', 'Yes', 'Sequential query number within database (1-30)', 'BIRD: —'],
    ['title', 'string', 'Yes', 'Short descriptive title (5-10 words)', 'Spider: question (partial)'],
    ['description', 'string (300-500 chars)', 'Yes', 'Natural language question a business user would ask. Must read as a genuine stakeholder question, NOT a technical summary.', 'BIRD: question / Spider: question'],
    ['purpose', 'string (200-400 chars)', 'Yes', 'WHY this analysis matters — the business decision or action it enables. Written from stakeholder perspective.', 'BIRD: evidence (extended)'],
    ['use_case', 'string (150-300 chars)', 'Yes', 'WHO uses this and WHEN — specific role, department, cadence, and trigger for running the query.', 'Semantic layer: context'],
    ['business_value', 'string (150-300 chars)', 'Yes', 'Quantifiable or qualifiable impact — revenue, cost savings, risk reduction, efficiency gain.', 'Semantic layer: ontology'],
    ['complexity', 'enum', 'Yes', 'One of: basic, intermediate, advanced, expert. Based on SQL features used.', 'BIRD: difficulty'],
    ['expected_output', 'object', 'Yes', 'Schema of result set: column names, data types, sample values, row count estimate.', 'Spider: db_id (implicit)'],
    ['sql', 'string', 'Yes', 'Production-quality SQL. Must execute without errors on the provided schema.', 'BIRD: SQL / Spider: query'],
    ['tags', 'array[string]', 'Recommended', 'Classification tags: domain area, SQL features used, business function.', 'BIRD: — (enhancement)'],
    ['evidence', 'string', 'Recommended', 'Domain knowledge or business rules needed to understand the query (BIRD-compatible).', 'BIRD: evidence'],
]

for i, f in enumerate(fields):
    row = 5 + i
    for j, val in enumerate(f):
        cell = ws2.cell(row=row, column=j+1, value=val)
        cell.font = black_font if j > 0 else black_bold
        cell.alignment = left_align
        cell.border = thin_border
        if j == 2 and val == 'Yes':
            cell.fill = green_fill
        elif j == 2:
            cell.fill = PatternFill('solid', fgColor='FFF2CC')

# Current vs recommended comparison
comp_start = 5 + len(fields) + 2
ws2.merge_cells(f'A{comp_start}:E{comp_start}')
ws2.cell(row=comp_start, column=1, value='CURRENT STATE vs RECOMMENDED STATE').font = section_font
ws2[f'A{comp_start}'].fill = light_fill

header_row(ws2, comp_start+1, ['Field', 'Well-Formed DBs (10)', 'Broken DBs (6)', 'Recommended Standard', 'Action Required'])

comparison = [
    ['description', '84-309 chars', '500 chars (truncated)', '300-500 chars, stakeholder voice', 'Rewrite all — too short or truncated'],
    ['purpose', 'Present', 'MISSING', 'Required 200-400 chars', 'Add to 6 broken DBs, expand in 10'],
    ['use_case', 'Present', 'MISSING', 'Required 150-300 chars', 'Add to 6 broken DBs, expand in 10'],
    ['business_value', 'Present', 'MISSING', 'Required 150-300 chars', 'Add to 6 broken DBs, expand in 10'],
    ['line_number', 'Absent', 'Present', 'REMOVE — not in spec', 'Delete from all DBs'],
    ['tags', 'Absent', 'Absent', 'Recommended — add', 'New field across all DBs'],
    ['evidence', 'Absent', 'Absent', 'Recommended (BIRD compat)', 'New field across all DBs'],
]

for i, c in enumerate(comparison):
    row = comp_start + 2 + i
    for j, val in enumerate(c):
        cell = ws2.cell(row=row, column=j+1, value=val)
        cell.font = black_font
        cell.alignment = left_align
        cell.border = thin_border
        if j == 2 and 'MISSING' in val:
            cell.fill = red_fill
            cell.font = red_font
        elif j == 4:
            cell.fill = yellow_fill

# ═══════════════════════════════════════════════════════
# SHEET 3: Rework Effort Estimate
# ═══════════════════════════════════════════════════════
ws3 = wb.create_sheet('Rework Estimate')
ws3.sheet_properties.tabColor = 'FF0000'

ws3.column_dimensions['A'].width = 40
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 20
ws3.column_dimensions['E'].width = 25

ws3.merge_cells('A1:E1')
ws3.cell(row=1, column=1, value='Rework & Remediation Cost for 16 Delivered Databases').font = title_font

ws3.merge_cells('A3:E3')
ws3.cell(row=3, column=1, value='REWORK EFFORT').font = section_font
ws3['A3'].fill = light_fill

header_row(ws3, 4, ['Task', 'Hours/Query', 'Queries Affected', 'Total Hours', 'Cost @ $100/hr'])

rework_tasks = [
    ['Rewrite descriptions (all 16 DBs) — stakeholder voice', 0.5, 452, None, None],
    ['Add purpose field (6 broken DBs)', 0.5, 180, None, None],
    ['Add use_case field (6 broken DBs)', 0.25, 180, None, None],
    ['Add business_value field (6 broken DBs)', 0.25, 180, None, None],
    ['Remove line_number, fix JSON structure', 0.1, 180, None, None],
    ['Expand purpose/use_case/biz_value (10 well-formed DBs)', 0.35, 272, None, None],
    ['Add tags field (all 16 DBs)', 0.15, 452, None, None],
    ['Add evidence/domain context (all 16 DBs)', 0.25, 452, None, None],
    ['QA & validation pass', 0.1, 452, None, None],
]

for i, t in enumerate(rework_tasks):
    row = 5 + i
    ws3.cell(row=row, column=1, value=t[0]).font = black_font
    ws3.cell(row=row, column=1).alignment = left_align
    ws3.cell(row=row, column=1).border = thin_border
    
    ws3.cell(row=row, column=2, value=t[1]).font = blue_font
    ws3.cell(row=row, column=2).number_format = decimal_fmt
    ws3.cell(row=row, column=2).border = thin_border
    ws3.cell(row=row, column=2).alignment = right_align
    
    ws3.cell(row=row, column=3, value=t[2]).font = blue_font
    ws3.cell(row=row, column=3).number_format = num_fmt
    ws3.cell(row=row, column=3).border = thin_border
    ws3.cell(row=row, column=3).alignment = right_align
    
    ws3.cell(row=row, column=4).value = f'=B{row}*C{row}'
    ws3.cell(row=row, column=4).font = black_font
    ws3.cell(row=row, column=4).number_format = decimal_fmt
    ws3.cell(row=row, column=4).border = thin_border
    ws3.cell(row=row, column=4).alignment = right_align
    
    ws3.cell(row=row, column=5).value = f'=D{row}*\'Program Economics\'!B{ri}'
    ws3.cell(row=row, column=5).font = black_font
    ws3.cell(row=row, column=5).number_format = dollar_fmt
    ws3.cell(row=row, column=5).border = thin_border
    ws3.cell(row=row, column=5).alignment = right_align

# Totals
tot_rw = 5 + len(rework_tasks)
ws3.cell(row=tot_rw, column=1, value='TOTAL REWORK').font = black_bold
ws3.cell(row=tot_rw, column=1).fill = light_fill
ws3.cell(row=tot_rw, column=1).border = thin_border
ws3.cell(row=tot_rw, column=4).value = f'=SUM(D5:D{tot_rw-1})'
ws3.cell(row=tot_rw, column=4).font = black_bold
ws3.cell(row=tot_rw, column=4).fill = light_fill
ws3.cell(row=tot_rw, column=4).number_format = decimal_fmt
ws3.cell(row=tot_rw, column=4).border = thin_border
ws3.cell(row=tot_rw, column=4).alignment = right_align
ws3.cell(row=tot_rw, column=5).value = f'=SUM(E5:E{tot_rw-1})'
ws3.cell(row=tot_rw, column=5).font = black_bold
ws3.cell(row=tot_rw, column=5).fill = light_fill
ws3.cell(row=tot_rw, column=5).number_format = dollar_fmt
ws3.cell(row=tot_rw, column=5).border = thin_border
ws3.cell(row=tot_rw, column=5).alignment = right_align

# Consulting rate equivalent
cr_row = tot_rw + 1
ws3.cell(row=cr_row, column=1, value='TOTAL REWORK @ CONSULTING RATE ($475/hr)').font = black_bold
ws3.cell(row=cr_row, column=1).border = thin_border
ws3.cell(row=cr_row, column=1).alignment = left_align
ws3.cell(row=cr_row, column=5).value = f'=D{tot_rw}*\'Program Economics\'!C{ri}'
ws3.cell(row=cr_row, column=5).font = red_font
ws3.cell(row=cr_row, column=5).number_format = dollar_fmt
ws3.cell(row=cr_row, column=5).border = thin_border
ws3.cell(row=cr_row, column=5).alignment = right_align

# Revenue earned for 16 DBs
rev16 = cr_row + 2
ws3.cell(row=rev16, column=1, value='Revenue Earned (16 DBs × $2K)').font = black_font
ws3.cell(row=rev16, column=1).border = thin_border
ws3.cell(row=rev16, column=1).alignment = left_align
ws3.cell(row=rev16, column=5).value = f'=\'Program Economics\'!B8*\'Program Economics\'!B{cv}'
ws3.cell(row=rev16, column=5).font = black_font
ws3.cell(row=rev16, column=5).number_format = dollar_fmt
ws3.cell(row=rev16, column=5).border = thin_border
ws3.cell(row=rev16, column=5).alignment = right_align

# Net position
net_row = rev16 + 1
ws3.cell(row=net_row, column=1, value='Net Position After Rework (Staffing Rate)').font = black_bold
ws3.cell(row=net_row, column=1).border = thin_border
ws3.cell(row=net_row, column=1).alignment = left_align
ws3.cell(row=net_row, column=5).value = f'=E{rev16}-E{tot_rw}'
ws3.cell(row=net_row, column=5).font = red_font
ws3.cell(row=net_row, column=5).number_format = dollar_fmt_neg
ws3.cell(row=net_row, column=5).border = thin_border
ws3.cell(row=net_row, column=5).alignment = right_align

out_path = '/sessions/great-intelligent-davinci/mnt/db/AQ_Program_Economics.xlsx'
wb.save(out_path)
print(f'Saved to {out_path}')
